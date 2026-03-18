from __future__ import annotations

import pathlib
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .models import NormalizedArticle


def export_raw_articles_excel(
    articles: List[NormalizedArticle],
    dest_path: pathlib.Path,
) -> pathlib.Path:
    """Export pre-cleaning articles to Excel.

    Each row contains: source vendor, publish time, original URL, title,
    signal type, source type.  These are raw fields captured before the
    signal-gate / role-specific-gate filtering.
    """
    dest_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Articles (Pre-Cleaning)"

    headers = [
        "来源厂商",
        "新闻时间",
        "原文链接",
        "标题",
        "信号类型",
        "来源类型",
        "区域",
    ]

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        bottom=Side(style="thin", color="D1D5DB"),
    )

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for idx, a in enumerate(articles, start=2):
        ws.cell(row=idx, column=1, value=a.company_or_firm_name)
        ws.cell(row=idx, column=2, value=a.published_at[:10] if a.published_at else "")
        ws.cell(row=idx, column=3, value=a.url)
        ws.cell(row=idx, column=4, value=a.title)
        ws.cell(row=idx, column=5, value=a.signal_type or "")
        ws.cell(row=idx, column=6, value=a.source_type or "")
        ws.cell(row=idx, column=7, value=a.region or "")
        for col in range(1, len(headers) + 1):
            ws.cell(row=idx, column=col).border = thin_border
            ws.cell(row=idx, column=col).alignment = Alignment(vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 10

    ws.auto_filter.ref = ws.dimensions

    wb.save(dest_path)
    return dest_path
